"""Stage 8: Excel export.

Produces output/operators_ma.xlsx with sheets:
  - Ranked         : all operators, one row each, with scores and key fields
  - Top 30 Targets : independents only (independence_score >= 0.7),
                     ranked by ma_fit_score
  - Groups         : consolidator + shared-signal rollups with member counts
  - Consolidators  : operators assigned to excluded consolidator groups
  - Summary        : headline counts

Uses openpyxl; falls back to a CSV dump if openpyxl not installed.
"""

from __future__ import annotations

import csv
import sqlite3
from pathlib import Path

from db import DB_PATH, connect, init_db

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "output"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_XLSX = OUT_DIR / "operators_ma.xlsx"


RANKED_SQL = """
SELECT
    s.rank,
    o.fta_no, o.name, o.legal_name, o.neq,
    o.city, o.postal, o.province,
    o.phone, o.email, o.website,
    o.president, o.delegue_votant,
    g.name AS group_name, g.kind AS group_kind, g.is_excluded,
    ROUND(s.ma_fit_score, 4) AS ma_fit_score,
    ROUND(s.size_score, 4)         AS size_score,
    ROUND(s.independence_score, 2) AS independence_score,
    ROUND(s.clarity_score, 2)      AS clarity_score,
    ROUND(s.succession_score, 2)   AS succession_score,
    (SELECT MAX(total)          FROM fleet    WHERE operator_id = o.id) AS fleet_total,
    (SELECT MAX(parking_area_m2) FROM facility WHERE operator_id = o.id) AS parking_m2,
    (SELECT COUNT(DISTINCT permit_type) FROM permits WHERE operator_id = o.id) AS n_permit_types,
    (SELECT COUNT(*) FROM ownership WHERE operator_id = o.id AND role = 'actionnaire') AS n_shareholders,
    s.rationale
FROM operators o
LEFT JOIN scores s ON s.operator_id = o.id
LEFT JOIN groups g ON o.group_id = g.id
ORDER BY s.rank
"""

TOP_TARGETS_SQL = """
SELECT
    s.rank,
    o.name, o.city, o.postal, o.phone, o.email, o.website,
    o.president, o.neq,
    ROUND(s.ma_fit_score, 4) ma_fit_score,
    ROUND(s.size_score, 4)   size_score,
    ROUND(s.clarity_score, 2) clarity_score,
    (SELECT MAX(total)          FROM fleet    WHERE operator_id = o.id) AS fleet_total,
    (SELECT MAX(parking_area_m2) FROM facility WHERE operator_id = o.id) AS parking_m2,
    (SELECT COUNT(DISTINCT permit_type) FROM permits WHERE operator_id = o.id) AS n_permit_types,
    g.name AS group_name,
    s.rationale
FROM operators o
JOIN scores s ON s.operator_id = o.id
LEFT JOIN groups g ON o.group_id = g.id
WHERE s.independence_score >= 0.7
ORDER BY s.ma_fit_score DESC
LIMIT 30
"""

GROUPS_SQL = """
SELECT g.name, g.kind, g.is_excluded,
       COUNT(o.id) AS n_members,
       GROUP_CONCAT(o.name, ' | ') AS members
FROM groups g
LEFT JOIN operators o ON o.group_id = g.id
GROUP BY g.id
ORDER BY g.is_excluded DESC, n_members DESC, g.name
"""

CONSOLIDATORS_SQL = """
SELECT g.name AS parent_group, o.name, o.city, o.president, o.email
FROM operators o JOIN groups g ON o.group_id = g.id
WHERE g.is_excluded = 1
ORDER BY g.name, o.name
"""


def _rows(conn: sqlite3.Connection, sql: str) -> tuple[list, list]:
    cur = conn.execute(sql)
    headers = [d[0] for d in cur.description]
    rows = [[r[h] for h in headers] for r in cur.fetchall()]
    return headers, rows


def export(db_path: Path = DB_PATH, out_path: Path = OUT_XLSX) -> Path:
    init_db(db_path)
    conn = connect(db_path)
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return _export_csv_fallback(conn)

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        counts = {
            "operators": conn.execute("SELECT COUNT(*) FROM operators").fetchone()[0],
            "groups": conn.execute("SELECT COUNT(*) FROM groups").fetchone()[0],
            "consolidator_subsidiaries": conn.execute(
                "SELECT COUNT(*) FROM operators o JOIN groups g "
                "ON o.group_id = g.id WHERE g.is_excluded = 1"
            ).fetchone()[0],
            "independents_candidates": conn.execute(
                "SELECT COUNT(*) FROM operators o LEFT JOIN groups g "
                "ON o.group_id = g.id WHERE g.is_excluded IS NULL OR g.is_excluded = 0"
            ).fetchone()[0],
            "with_neq":     conn.execute("SELECT COUNT(*) FROM operators WHERE neq IS NOT NULL").fetchone()[0],
            "with_lat":     conn.execute("SELECT COUNT(*) FROM operators WHERE lat IS NOT NULL").fetchone()[0],
            "with_permits": conn.execute("SELECT COUNT(DISTINCT operator_id) FROM permits").fetchone()[0],
            "with_facility":conn.execute("SELECT COUNT(DISTINCT operator_id) FROM facility").fetchone()[0],
            "scored":       conn.execute("SELECT COUNT(*) FROM scores").fetchone()[0],
        }
        ws_summary.append(["Metric", "Value"])
        for k, v in counts.items():
            ws_summary.append([k, v])
        for c in ws_summary[1]:
            c.font = Font(bold=True)

        for title, sql in [
            ("Ranked",        RANKED_SQL),
            ("Top 30 Targets", TOP_TARGETS_SQL),
            ("Groups",        GROUPS_SQL),
            ("Consolidators", CONSOLIDATORS_SQL),
        ]:
            ws = wb.create_sheet(title)
            headers, rows = _rows(conn, sql)
            ws.append(headers)
            for h_cell in ws[1]:
                h_cell.font = Font(bold=True)
                h_cell.fill = PatternFill("solid", fgColor="DDDDDD")
            for row in rows:
                ws.append(row)
            ws.freeze_panes = "A2"
            for col_idx, _h in enumerate(headers, 1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                width = max(12, min(60, max(
                    (len(str(r[col_idx - 1])) for r in rows[:200]), default=12
                ) + 2))
                ws.column_dimensions[col_letter].width = width

        wb.save(out_path)
        return out_path
    finally:
        conn.close()


def _export_csv_fallback(conn: sqlite3.Connection) -> Path:
    print("openpyxl not installed — writing CSVs instead")
    for name, sql in [("ranked", RANKED_SQL), ("top30", TOP_TARGETS_SQL),
                      ("groups", GROUPS_SQL), ("consolidators", CONSOLIDATORS_SQL)]:
        p = OUT_DIR / f"{name}.csv"
        headers, rows = _rows(conn, sql)
        with p.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(rows)
        print(f"  wrote {p}  ({len(rows)} rows)")
    return OUT_DIR


if __name__ == "__main__":
    p = export()
    print(f"Wrote {p}")
